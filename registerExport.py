import ctypes, sys, tkinter as tk, openpyxl
from tkinter import filedialog, messagebox
from openpyxl.formatting.rule import Rule
from copy import copy

# Funció amagar consola (programa windowless)
def windowless():
    kernel32 = ctypes.WinDLL('kernel32')
    user32 = ctypes.WinDLL('user32')
    SW_HIDE = 0
    hWnd = kernel32.GetConsoleWindow()
    if hWnd:
        user32.ShowWindow(hWnd, SW_HIDE)

# Amagar la consola a l'inici de l'script
if sys.platform == 'win32':
    windowless()

def selecciona_fitxer(tipus):
    fitxer = filedialog.askopenfilename(filetypes=[("Fitxers Excel", "*.xlsx;*.xls")])
    if tipus == "base":
        fitxerBase.delete(0, tk.END)
        fitxerBase.insert(0, fitxer)
    elif tipus == "Dades":
        fitxerDades.delete(0, tk.END)
        fitxerDades.insert(0, fitxer)

def processa_fitxers():
    ruta_base = fitxerBase.get()
    ruta_Dades = fitxerDades.get()

    if ruta_base == "" or ruta_Dades == "":
        messagebox.showerror("Error", "Selecciona dos fitxers.")
        return

    try:
        doc_base = openpyxl.load_workbook(ruta_base)
        doc_dades = openpyxl.load_workbook(ruta_Dades)

        fullPlantilla = doc_base['estadistica']

        delete_sheet = [fulla for fulla in doc_dades.sheetnames if fulla.startswith('Informe dades recopilació')]
        for fulla in delete_sheet:
            del doc_dades[fulla]

        fullNou = doc_dades.create_sheet(title="Informe dades recopilació")

        for fila in fullPlantilla.iter_rows():
            for cel in fila:
                nova_cel = fullNou[cel.coordinate]
                nova_cel.value = cel.value
                if cel.has_style:
                    nova_cel.font = copy(cel.font)
                    nova_cel.border = copy(cel.border)
                    nova_cel.fill = copy(cel.fill)
                    nova_cel.number_format = copy(cel.number_format)
                    nova_cel.protection = copy(cel.protection)
                    nova_cel.alignment = copy(cel.alignment)

        for cel_mer in fullPlantilla.merged_cells.ranges:
            fullNou.merge_cells(str(cel_mer))
            min_col, min_row, max_col, max_row = cel_mer.bounds
            cel_superior_esquerra = fullPlantilla.cell(row=min_row, column=min_col)
            nova_cel_superior_esquerra = fullNou.cell(row=min_row, column=min_col)
            nova_cel_superior_esquerra.value = cel_superior_esquerra.value
            if cel_superior_esquerra.has_style:
                nova_cel_superior_esquerra.font = copy(cel_superior_esquerra.font)
                nova_cel_superior_esquerra.border = copy(cel_superior_esquerra.border)
                nova_cel_superior_esquerra.fill = copy(cel_superior_esquerra.fill)
                nova_cel_superior_esquerra.number_format = copy(cel_superior_esquerra.number_format)
                nova_cel_superior_esquerra.protection = copy(cel_superior_esquerra.protection)
                nova_cel_superior_esquerra.alignment = copy(cel_superior_esquerra.alignment)

        for columna in fullPlantilla.columns:
            columna_letra = columna[0].column_letter
            fullNou.column_dimensions[columna_letra].width = fullPlantilla.column_dimensions[columna_letra].width

        for range_string, regles in fullPlantilla.conditional_formatting._cf_rules.items():
            for regla in regles:
                if isinstance(regla, Rule):
                    nova_regla = Rule(
                        type=regla.type,
                        dxf=copy(regla.dxf),
                        formula=regla.formula,
                        stopIfTrue=regla.stopIfTrue
                    )
                    fullNou.conditional_formatting.add(range_string, nova_regla)

        ruta_guardat = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Fitxers Excel", "*.xlsx")])
        if not ruta_guardat:
            return

        doc_dades.save(ruta_guardat)
        messagebox.showinfo("Èxit", "Dades copiades correctament.")
        finestra.destroy()
    except FileNotFoundError:
        messagebox.showerror("Error", "Fitxer no trobat.")
    except PermissionError:
        messagebox.showerror("Error", "Permís denegat per desar el fitxer.")
    except Exception as e:
        messagebox.showerror("Error", f"S'ha produït un error: {str(e)}")

finestra = tk.Tk()
finestra.title("Còpia de Dades i Format d'Excel")
finestra.geometry("400x250")

baseWindow = tk.Label(finestra, text="Fitxer Base:")
baseWindow.pack(pady=10)
fitxerBase = tk.Entry(finestra, width=50)
fitxerBase.pack()
bt_baseWindow = tk.Button(finestra, text="Selecciona Base", command=lambda: selecciona_fitxer("base"))
bt_baseWindow.pack(pady=5)

dadesWindow = tk.Label(finestra, text="Fitxer Dades:")
dadesWindow.pack(pady=10)
fitxerDades = tk.Entry(finestra, width=50)
fitxerDades.pack()
bt_dadesWindow = tk.Button(finestra, text="Selecciona Dades", command=lambda: selecciona_fitxer("Dades"))
bt_dadesWindow.pack(pady=5)

bt_make = tk.Button(finestra, text="Processa Fitxers", command=processa_fitxers)
bt_make.pack(pady=20)

finestra.mainloop()
